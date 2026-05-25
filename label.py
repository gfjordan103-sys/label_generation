import streamlit as st
import openpyxl
from openpyxl.styles import Font, Border, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

# ==========================================
# 核心邏輯：20欄一頁完美Cycle，欄寬格式完全同步樣版
# ==========================================
def generate_exact_printing_cycle_labels(detail_wb, template_wb):
    ws_detail = detail_wb.worksheets[0]
    ws_template = template_wb.worksheets[0]
    
    # 1. 解析出貨明細
    board_data = []
    board_dict = {}
    current_board = None
    
    for row in range(3, ws_detail.max_row + 1):
        board_val = ws_detail.cell(row=row, column=1).value  # A欄
        box_val = ws_detail.cell(row=row, column=7).value    # G欄
        order_val = ws_detail.cell(row=row, column=8).value  # H欄
        
        if board_val and str(board_val).strip():
            current_board = str(board_val).strip()
            if current_board not in board_dict:
                board_dict[current_board] = {'total_boxes': 0, 'order_no': ''}
                board_data.append(current_board)
        
        if current_board and box_val is not None:
            try:
                board_dict[current_board]['total_boxes'] += int(float(box_val))
            except ValueError:
                pass
            if order_val and not board_dict[current_board]['order_no']:
                board_dict[current_board]['order_no'] = str(order_val).split('.')[0]

    # 2. 建立全新的列印活頁簿
    output_wb = openpyxl.Workbook()
    ws_output = output_wb.active
    ws_output.title = "標籤自動排版列印表"
    
    # 一比一讀取並緩存樣版 A(1) 到 T(20) 的【精確欄寬】與【每欄儲存格樣式】
    template_widths = {}
    template_styles = {} # 記錄每一欄的格線與格式特徵
    
    for c_idx in range(1, 21):
        col_letter = get_column_letter(c_idx)
        # A. 記錄欄寬
        if col_letter in ws_template.column_dimensions and ws_template.column_dimensions[col_letter].width is not None:
            template_widths[c_idx] = ws_template.column_dimensions[col_letter].width
        else:
            template_widths[c_idx] = 10.0 # 預設防呆寬度
            
        # B. 記錄該欄第3列的基礎標籤格子樣式 (做為格式複製來源)
        ref_cell = ws_template.cell(row=3, column=c_idx)
        template_styles[c_idx] = {
            'font': Font(**ref_cell.font.__dict__) if ref_cell.font else None,
            'border': Border(**ref_cell.border.__dict__) if ref_cell.border else None,
            'fill': PatternFill(**ref_cell.fill.__dict__) if ref_cell.fill else None,
            'alignment': Alignment(**ref_cell.alignment.__dict__) if ref_cell.alignment else None,
            'number_format': ref_cell.number_format
        }

    # 讀取樣版第 2 列 A 欄的板號專用字體
    ref_hdr_cell = ws_template.cell(row=2, column=1)
    hdr_font = Font(**ref_hdr_cell.font.__dict__) if ref_hdr_cell.font else Font(name='Arial', size=12, bold=True)

    # 3. 排版引擎變數
    # 根據樣版設計：第一頁 A~T 欄共包含 7 組標籤。每組橫向佔用 2 欄內容，並由間隔欄隔開。
    # 標籤在單頁內的相對起始欄索引分別為：1(A), 4(D), 7(G), 10(J), 13(M), 16(P), 19(S)
    label_group_offsets = [1, 4, 7, 10, 13, 16, 19]
    current_group_global_index = 0  
    MAX_LABELS_PER_COL = 32 # 限制：第 1 列空白標頭，第 2 列板號，下方最多放 32 列標籤（剛好限高34列）

    for board_name in board_data:
        info = board_dict[board_name]
        boxes_count = info['total_boxes']
        order_no = info['order_no']
        
        if boxes_count <= 0:
            continue
            
        remaining_boxes = boxes_count
        
        while remaining_boxes > 0:
            current_col_boxes = min(remaining_boxes, MAX_LABELS_PER_COL)
            
            # 計算這組標籤在 Cycle 規則下的絕對起始欄位
            page_number = current_group_global_index // 7   
            group_in_page = current_group_global_index % 7  
            start_col_idx = (page_number * 20) + label_group_offsets[group_in_page]
            
            # A. 寫入第 2 列：板號名稱 (連續填寫一組的 2 個欄位，確保網格同步)
            for c_offset in range(2):
                target_c = start_col_idx + c_offset
                tpl_idx = label_group_offsets[group_in_page] + c_offset # 對應第1頁的哪一欄
                style = template_styles[tpl_idx]
                
                h_cell = ws_output.cell(row=2, column=target_c)
                h_cell.font = hdr_font
                if style['border']: h_cell.border = style['border']
                if style['fill']: h_cell.fill = style['fill']
                if style['alignment']: h_cell.alignment = style['alignment']
                
                if c_offset == 0:
                    h_cell.value = board_name

            # B. 從第 3 列到第 34 列：寫入訂單號碼標籤，並複製對應欄位的格線格式
            for i in range(1, current_col_boxes + 1):
                target_row = 2 + i  
                for c_offset in range(2):
                    target_c = start_col_idx + c_offset
                    tpl_idx = label_group_offsets[group_in_page] + c_offset # 對應第1頁的哪一欄
                    style = template_styles[tpl_idx]
                    
                    cell = ws_output.cell(row=target_row, column=target_c)
                    if style['font']: cell.font = style['font']
                    if style['border']: cell.border = style['border']
                    if style['fill']: cell.fill = style['fill']
                    if style['alignment']: cell.alignment = style['alignment']
                    cell.number_format = style['number_format']
                    
                    if c_offset == 0:
                        cell.value = order_no
            
            remaining_boxes -= current_col_boxes
            current_group_global_index += 1

    # 【核心修正】整張報表生成後，強行實施 20 欄 Cycle 模數外推，強制重設右邊所有新增欄位的寬度！
    total_used_cols = ((current_group_global_index - 1) // 7 + 1) * 20
    for c_idx in range(1, total_used_cols + 1):
        c_let = get_column_letter(c_idx)
        # 20欄模數循環公式：不管是第幾頁，第 c_idx 欄的寬度永遠跟隨第1頁 (1~20) 對應欄位
        tpl_idx = ((c_idx - 1) % 20) + 1
        ws_output.column_dimensions[c_let].width = template_widths[tpl_idx]

    return output_wb

# ==========================================
# Streamlit 純網頁 UI 介面
# ==========================================
st.set_page_config(page_title="印刷標籤自動排版系統", page_icon="🖨️")
st.title("🖨️ 印刷標籤自動排版系統 (列寬完美同步版)")
st.write("【完美規範：A~T欄為一頁橫向Cycle，限高 34 列】。所有新增欄寬一比一跟隨樣版循環，C, F, I 等窄欄寬不走鐘。")

file_detail = st.file_uploader("1. 請上傳【出貨明細】Excel 檔案 (.xlsx)", type=["xlsx"], key="p_detail")
file_template = st.file_uploader("2. 請上傳【訂單號碼】標籤樣版 Excel 檔案 (.xlsx)", type=["xlsx"], key="p_template")

if file_detail and file_template:
    st.success("📊 雙檔案載入成功！")
    if st.button("🚀 執行 20 欄 Cycle 印刷排版", type="primary"):
        with st.spinner("正在進行精密橫向分頁對位與欄寬強制同步..."):
            try:
                wb_d = openpyxl.load_workbook(file_detail)
                wb_t = openpyxl.load_workbook(file_template)
                
                # 執行印刷分頁循環邏輯
                result_wb = generate_printing_page_labels(wb_d, wb_t)
                
                excel_buffer = io.BytesIO()
                result_wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.success("🎉 排版成功！欄寬已完美實現 20 欄橫向 Cycle 複製，請點選下方下載。")
                st.download_button(
                    label="📥 下載印刷對位標籤結果",
                    data=excel_buffer,
                    file_name="20欄Cycle_印刷標籤結果.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"❌ 錯誤：{e}")
